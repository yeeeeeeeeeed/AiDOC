"""엑셀 생성 서비스"""

import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def tables_to_excel(tables: list) -> bytes:
    """테이블 리스트 → 엑셀 바이트"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    header_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=10, color="7A4010")
    title_fill = PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid")

    for idx, table in enumerate(tables):
        title = table.get("title", f"표{idx+1}")
        safe_title = "".join(c for c in title if c not in r'[]:*?/\\')[:31]
        if not safe_title:
            safe_title = f"Sheet{idx+1}"
        base = safe_title
        cnt = 1
        while safe_title in wb.sheetnames:
            safe_title = f"{base[:28]}({cnt})"
            cnt += 1

        ws = wb.create_sheet(safe_title)
        headers = table.get("headers", [])
        rows = table.get("rows", [])

        # A열: 테이블명
        cell = ws.cell(row=1, column=1, value="테이블명")
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

        for c, h in enumerate(headers, 2):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center

        for r, row_data in enumerate(rows, 2):
            cell = ws.cell(row=r, column=1, value=title)
            cell.font = title_font
            cell.fill = title_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            for c, val in enumerate(row_data, 2):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = border
                try:
                    num = float(str(val).replace(",", ""))
                    cell.value = num
                    cell.number_format = "#,##0"
                except (ValueError, TypeError):
                    pass

        ws.column_dimensions["A"].width = 30
        total_cols = len(headers) + 1
        for c in range(2, total_cols + 1):
            max_len = max(
                (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, len(rows) + 2)),
                default=8,
            )
            ws.column_dimensions[get_column_letter(c)].width = max(10, min(max_len + 4, 40))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
